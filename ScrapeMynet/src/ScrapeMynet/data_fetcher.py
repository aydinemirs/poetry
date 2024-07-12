import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from colorama import Fore, Style  

def main():
    url = 'https://finans.mynet.com/borsa/hisseler/'

    response = requests.get(url)
    if response.status_code == 200:
        page_content = response.content
    else:
        print(f"Error: {response.status_code}")
        return

    soup = BeautifulSoup(page_content, 'html.parser')

    tables = soup.find_all('table')
    if not tables:
        print("Tablo bulunamadı. HTML yapısını kontrol edin.")
        return

    data = []
    for i, table in enumerate(tables):
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all('td')
            row_data = []
            for cell in cells:
                text = cell.get_text(strip=True)
                row_data.append(text)  
            if row_data:
                data.append(row_data)

    df = pd.DataFrame(data)

    excel_file = "hisseler.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']  

        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = int((max_length + 2) * 1.25) 
            worksheet.column_dimensions[column].width = adjusted_width
        
    print(f"{Fore.GREEN}Veriler '{excel_file}' dosyasına yazdırıldı.{Style.RESET_ALL}")

    while True:
        hisse_adi = input(f"{Fore.WHITE}Lütfen bir hisse adı girin (Çıkmak için 'exit' yazın): ").strip().upper()

        if hisse_adi == 'EXIT':
            break
        
        wb = load_workbook(excel_file)
        sheet = wb.active

        found = False
        for row in sheet.iter_rows(values_only=True):
            if hisse_adi in str(row[0]).upper(): 
                print(f"{Fore.CYAN}Hisse Adı: {Style.BRIGHT}{Fore.WHITE}{row[0]}")
                print(f"{Fore.CYAN}Son Fiyat: {Style.BRIGHT}{Fore.WHITE}{row[2]}")
                print(f"{Fore.CYAN}Değişim Yüzde: {Style.BRIGHT}{Fore.WHITE}{row[3]}")
                print(f"{Fore.CYAN}Hacim(TL): {Style.BRIGHT}{Fore.WHITE}{row[4]}")
                print(f"{Fore.CYAN}Saat: {Style.BRIGHT}{Fore.WHITE}{row[5]}")
                found = True
                break

        if not found:
            print(f"{Fore.RED}Hisse adı '{hisse_adi}' bulunamadı.")
            print(f"{Fore.WHITE}............")

        wb.close()

if __name__ == "__main__":
    main()
